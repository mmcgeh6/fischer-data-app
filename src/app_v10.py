"""
Fischer Energy Partners - Data Processing Application V9
Enhanced Flagging, Multi-Tab Excel Support, and File Archiving
- Modified stale flag logic (3+ consecutive non-zero values)
- New Zero_Value_Flag column (Clear/Single/Repeated)
- Multi-tab Excel file support with column selection
- File archiving system with building name
- Fischer Energy branding and styling
"""

import streamlit as st
import pandas as pd
from pathlib import Path
from datetime import datetime, timedelta
import warnings
import json
import os
import io
import base64
from dotenv import load_dotenv
from anthropic import Anthropic
from concurrent.futures import ThreadPoolExecutor, as_completed
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, numbers
from openpyxl.utils.dataframe import dataframe_to_rows
import shutil

# Add src directory to path for imports
sys.path.insert(0, str(Path(__file__).parent))
from timestamp_normalizer import format_timestamp_mdy_hms, detect_timestamp_format

warnings.filterwarnings('ignore')

# Load environment variables
load_dotenv()

# Page configuration
st.set_page_config(
    page_title="Fischer Data Processing - V10",
    page_icon="📊",
    layout="wide"
)

# Initialize session state
if 'uploaded_files' not in st.session_state:
    st.session_state.uploaded_files = {}
if 'file_configs' not in st.session_state:
    st.session_state.file_configs = {}
if 'combined_df' not in st.session_state:
    st.session_state.combined_df = None
if 'resampled_df' not in st.session_state:
    st.session_state.resampled_df = None
if 'resampling_stats' not in st.session_state:
    st.session_state.resampling_stats = {}
if 'inexact_cells' not in st.session_state:
    st.session_state.inexact_cells = {}  # Track which cells are inexact for coloring
if 'ai_debug_log' not in st.session_state:
    st.session_state.ai_debug_log = []
if 'ai_analysis_complete' not in st.session_state:
    st.session_state.ai_analysis_complete = False
if 'building_name' not in st.session_state:
    st.session_state.building_name = ""
if 'archive_path' not in st.session_state:
    st.session_state.archive_path = ""


def add_logo():
    """Add Fischer Energy logo to top-left with title."""
    logo_path = Path(__file__).parent.parent / "assets" / "fischer background clear (1).png"

    if logo_path.exists():
        with open(logo_path, "rb") as f:
            logo_data = base64.b64encode(f.read()).decode()

        st.markdown(
            f"""
            <div style="display: flex; align-items: center; padding: 10px 0; margin-bottom: 20px;">
                <img src="data:image/png;base64,{logo_data}"
                     style="height: 80px; margin-right: 20px;">
                <h1 style="color: #24b3aa; margin: 0; font-size: 2.5rem;">Fischer Data Processor V10</h1>
            </div>
            """,
            unsafe_allow_html=True
        )
    else:
        st.title("🔧 Fischer Data Processor V10")


def inject_custom_css():
    """Inject custom CSS for Fischer Energy branding."""
    st.markdown(
        """
        <style>
        /* Button styling */
        .stButton > button {
            background-color: #24b3aa;
            color: #FFFFFF;
            border: none;
            border-radius: 4px;
            padding: 0.5rem 1rem;
            font-weight: 500;
        }
        .stButton > button:hover {
            background-color: #1e9b93;
        }

        /* Primary button styling */
        .stButton > button[kind="primary"] {
            background-color: #24b3aa;
            color: #FFFFFF;
        }
        .stButton > button[kind="primary"]:hover {
            background-color: #1e9b93;
        }

        /* Header styling */
        h1, h2, h3 {
            color: #24b3aa;
        }

        /* Expander headers */
        .streamlit-expanderHeader {
            background-color: #f0f9f8;
            border-left: 3px solid #24b3aa;
            font-weight: 500;
        }

        /* Success messages */
        .stSuccess {
            background-color: #e6f7f6;
            border-left: 4px solid #24b3aa;
        }

        /* Info messages */
        .stInfo {
            background-color: #f0f9f8;
            border-left: 4px solid #24b3aa;
        }

        /* Dataframe headers */
        .dataframe thead tr th {
            background-color: #24b3aa !important;
            color: #FFFFFF !important;
        }

        /* Sidebar styling */
        section[data-testid="stSidebar"] {
            background-color: #f0f9f8;
        }

        /* Download button */
        .stDownloadButton > button {
            background-color: #24b3aa;
            color: #FFFFFF;
        }
        .stDownloadButton > button:hover {
            background-color: #1e9b93;
        }

        /* Tab container styling - adds teal border around entire tab interface */
        .stTabs {
            border: 2px solid #24b3aa;
            border-radius: 8px;
            padding: 16px;
            background-color: #ffffff;
            margin: 16px 0;
        }

        /* Tab list styling - separator between tabs */
        .stTabs [data-baseweb="tab-list"] {
            gap: 8px;
            border-bottom: 2px solid #e0e0e0;
            padding-bottom: 8px;
            margin-bottom: 16px;
        }

        /* Individual tab styling */
        .stTabs [data-baseweb="tab"] {
            border-right: 1px solid #e0e0e0;
            padding-right: 16px;
            padding-left: 16px;
        }

        /* Remove border from last tab */
        .stTabs [data-baseweb="tab"]:last-child {
            border-right: none;
        }

        /* Active tab styling */
        .stTabs [aria-selected="true"] {
            border-bottom: 3px solid #24b3aa;
            color: #24b3aa;
            font-weight: 600;
        }

        /* Tab panel content styling */
        .stTabs [data-baseweb="tab-panel"] {
            padding: 16px;
            background-color: #fafafa;
            border-radius: 4px;
            margin-top: 8px;
        }

        /* Nested tabs styling (for sheet-level tabs inside file-level tabs) */
        .stTabs .stTabs {
            border: 1px solid #d0d0d0;
            border-radius: 4px;
            padding: 12px;
            background-color: #ffffff;
            margin-top: 8px;
        }

        .stTabs .stTabs [data-baseweb="tab-list"] {
            border-bottom: 1px solid #d0d0d0;
        }
        </style>
        """,
        unsafe_allow_html=True
    )


# Apply branding
add_logo()
inject_custom_css()


def read_raw_lines(file_path, num_lines=15):
    """Read first N lines of a file as raw text. For Excel files, convert to CSV-like format."""
    lines = []

    # Check if it's an Excel file
    if str(file_path).lower().endswith(('.xlsx', '.xls')):
        try:
            df = pd.read_excel(file_path, header=None, nrows=num_lines,
                             dtype=str, keep_default_na=False)
            csv_buffer = io.StringIO()
            df.to_csv(csv_buffer, index=False, header=False)
            csv_text = csv_buffer.getvalue()
            lines = csv_text.strip().split('\n')
            return lines[:num_lines]
        except Exception as e:
            return [f"Error reading Excel file: {str(e)}"]

    # For CSV/text files
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            for i, line in enumerate(f):
                if i >= num_lines:
                    break
                lines.append(line.rstrip())
    except UnicodeDecodeError:
        with open(file_path, 'r', encoding='latin-1') as f:
            for i, line in enumerate(f):
                if i >= num_lines:
                    break
                lines.append(line.rstrip())
    return lines


def detect_file_type(file_path):
    """
    Detect if file is CSV or Excel, and if Excel has multiple tabs.

    Returns:
        Tuple of (file_type, sheet_names)
        - file_type: 'excel_multi_tab', 'excel_single_tab', or 'csv'
        - sheet_names: List of tab names (None for CSV files)
    """
    if str(file_path).lower().endswith(('.xlsx', '.xls')):
        try:
            xl_file = pd.ExcelFile(file_path)
            sheet_names = xl_file.sheet_names
            num_tabs = len(sheet_names)

            if num_tabs > 1:
                return 'excel_multi_tab', sheet_names
            else:
                return 'excel_single_tab', sheet_names
        except Exception as e:
            return 'csv', None
    else:
        return 'csv', None


def read_tab_raw_lines(file_path, sheet_name, num_lines=15):
    """Read first N lines from a specific Excel tab."""
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=num_lines,
                         dtype=str, keep_default_na=False)
        csv_buffer = io.StringIO()
        df.to_csv(csv_buffer, index=False, header=False)
        csv_text = csv_buffer.getvalue()
        lines = csv_text.strip().split('\n')
        return lines[:num_lines]
    except Exception as e:
        return [f"Error reading tab {sheet_name}: {str(e)}"]


def build_ai_prompt(file_name, text_sample):
    """Build the AI prompt for column detection."""
    return f"""Analyze the first 15 lines of this CSV/Excel file sample to determine the column configuration for data processing.

File name: {file_name or 'unknown'}

Raw file content (first 15 lines):
{text_sample}

Based on this data, identify:
1. The delimiter used (comma, tab, semicolon, etc.)
2. Which row index (0-based) contains the column headers and is the START of data
3. Which column index (0-based) contains the date/timestamp
4. Which column index (0-based) contains the sensor value/reading
5. A sensor name (extract from metadata or suggest based on filename)

Return ONLY a JSON object in this exact format with no additional text:
{{
  "delimiter": ",",
  "start_row": 1,
  "date_column": 0,
  "value_column": 2,
  "sensor_name": "sensor name here"
}}

Rules:
- All indices MUST be 0-based
- start_row is where column headers are located (data begins on start_row+1)
- If a column doesn't exist, use -1 as the value
- Return only valid JSON with no explanations or additional text"""


def build_multi_tab_ai_prompt(file_name, tabs_data):
    """
    Build AI prompt for multi-tab Excel file analysis.

    Args:
        file_name: Name of the Excel file
        tabs_data: Dict of {tab_name: text_sample} for each tab

    Returns:
        Formatted prompt string for AI analysis
    """
    tabs_text = ""
    for tab_name, text_sample in tabs_data.items():
        tabs_text += f"\n\n=== TAB: {tab_name} ===\n{text_sample}\n"

    return f"""Analyze this multi-tab Excel file to determine column configurations for data processing.

File name: {file_name or 'unknown'}

{tabs_text}

For EACH tab, identify:
1. Which row index (0-based) contains the column headers (start_row)
2. Which column index (0-based) contains the date/timestamp
3. Which column indices (0-based) contain sensor values/readings - MULTIPLE columns are expected per tab
4. The names of the value columns (extract from headers or suggest descriptive names)

Return ONLY a JSON object in this exact format with no additional text:
{{
  "tabs": [
    {{
      "tab_name": "AC12-1",
      "start_row": 1,
      "date_column": 0,
      "value_columns": [2, 3, 4],
      "column_names": ["Return Air Temp", "Supply Air Temp", "Fan Status"]
    }},
    {{
      "tab_name": "AC12-2",
      "start_row": 1,
      "date_column": 0,
      "value_columns": [2, 3],
      "column_names": ["Return Air Temp", "Supply Air Temp"]
    }}
  ]
}}

Rules:
- All indices MUST be 0-based
- Analyze ALL tabs provided above
- value_columns should be a LIST of column indices (can be multiple per tab)
- column_names should match the order of value_columns
- Return only valid JSON with no explanations"""


def call_claude_api(prompt, api_key):
    """Call Claude API and return the response text."""
    client = Anthropic(api_key=api_key)

    try:
        response = client.messages.create(
            model="claude-sonnet-4-5-20250929",
            max_tokens=15000,
            temperature=0.6,
            messages=[{"role": "user", "content": prompt}]
        )

        # Extract text from response
        response_text = response.content[0].text.strip()
        return response_text
    except Exception as e:
        raise RuntimeError(f"Claude API call failed: {str(e)}")


def analyze_single_file(file_name, file_path, api_key):
    """
    Analyze a single CSV or single-tab Excel file with AI.

    Returns tuple of (config, debug_entry)
    """
    debug_entry = {
        'file_name': file_name,
        'timestamp': datetime.now().strftime('%H:%M:%S'),
        'request': None,
        'response': None,
        'error': None,
        'success': False
    }

    try:
        # Read first 15 lines
        raw_lines = read_raw_lines(file_path, num_lines=15)
        text_sample = "\n".join(raw_lines)

        # Build prompt
        prompt = build_ai_prompt(file_name, text_sample)

        # Store request
        debug_entry['request'] = {
            'model': 'claude-sonnet-4-5-20250929',
            'max_tokens': 15000,
            'temperature': 0.6,
            'prompt_length': len(prompt),
            'prompt_preview': prompt[:500] + '...' if len(prompt) > 500 else prompt,
            'raw_text_lines': raw_lines[:5]
        }

        # Call API
        response_text = call_claude_api(prompt, api_key)

        # Store response
        debug_entry['response'] = {
            'raw_text': response_text,
            'response_length': len(response_text)
        }

        # Parse JSON
        json_start = response_text.find('{')
        json_end = response_text.rfind('}') + 1
        if json_start != -1 and json_end > json_start:
            json_str = response_text[json_start:json_end]
            config = json.loads(json_str)

            debug_entry['response']['parsed_json'] = config
            debug_entry['success'] = True

            return config, debug_entry
        else:
            debug_entry['error'] = "Could not find JSON in response"
            return None, debug_entry

    except json.JSONDecodeError as e:
        debug_entry['error'] = f"JSON decode error: {str(e)}"
        return None, debug_entry
    except Exception as e:
        debug_entry['error'] = f"Error: {str(e)}"
        return None, debug_entry


def analyze_multi_tab_file(file_name, file_path, sheet_names, api_key):
    """
    Analyze a multi-tab Excel file with AI.

    Returns tuple of (config, debug_entry)
    Config format: {
        "file_type": "excel_multi_tab",
        "tabs": {
            "Tab1": {
                "start_row": 1,
                "date_column": 0,
                "available_columns": [2, 3, 4],
                "column_names": ["Col1", "Col2", "Col3"],
                "selected_columns": [2, 3, 4]  # Initially all selected
            }
        }
    }
    """
    debug_entry = {
        'file_name': file_name,
        'timestamp': datetime.now().strftime('%H:%M:%S'),
        'request': None,
        'response': None,
        'error': None,
        'success': False
    }

    try:
        # Read first 15 lines from each tab
        tabs_data = {}
        for sheet_name in sheet_names:
            raw_lines = read_tab_raw_lines(file_path, sheet_name, num_lines=15)
            tabs_data[sheet_name] = "\n".join(raw_lines)

        # Build multi-tab prompt
        prompt = build_multi_tab_ai_prompt(file_name, tabs_data)

        # Store request
        debug_entry['request'] = {
            'model': 'claude-sonnet-4-5-20250929',
            'max_tokens': 15000,  # Increased for multiple tabs
            'temperature': 0.6,
            'prompt_length': len(prompt),
            'prompt_preview': prompt[:500] + '...' if len(prompt) > 500 else prompt,
            'tabs_analyzed': list(sheet_names)
        }

        # Call API
        response_text = call_claude_api(prompt, api_key)

        # Store response
        debug_entry['response'] = {
            'raw_text': response_text,
            'response_length': len(response_text)
        }

        # Parse JSON
        json_start = response_text.find('{')
        json_end = response_text.rfind('}') + 1
        if json_start != -1 and json_end > json_start:
            json_str = response_text[json_start:json_end]
            ai_response = json.loads(json_str)

            # Convert AI response to our internal format
            config = {
                "file_type": "excel_multi_tab",
                "tabs": {}
            }

            for tab_data in ai_response.get('tabs', []):
                tab_name = tab_data['tab_name']
                value_cols = tab_data['value_columns']

                config['tabs'][tab_name] = {
                    'start_row': tab_data['start_row'],
                    'date_column': tab_data['date_column'],
                    'available_columns': value_cols,
                    'column_names': tab_data['column_names'],
                    'selected_columns': value_cols.copy()  # Initially all selected
                }

            debug_entry['response']['parsed_json'] = config
            debug_entry['success'] = True

            return config, debug_entry
        else:
            debug_entry['error'] = "Could not find JSON in response"
            return None, debug_entry

    except json.JSONDecodeError as e:
        debug_entry['error'] = f"JSON decode error: {str(e)}"
        return None, debug_entry
    except Exception as e:
        debug_entry['error'] = f"Error: {str(e)}"
        return None, debug_entry


def analyze_file_with_detection(file_name, file_path, api_key):
    """
    Detect file type and analyze accordingly.

    Returns tuple of (config, debug_entry)
    """
    # Detect file type
    file_type, sheet_names = detect_file_type(file_path)

    if file_type == 'excel_multi_tab':
        # Analyze multi-tab Excel file
        config, debug_entry = analyze_multi_tab_file(file_name, file_path, sheet_names, api_key)
        if config:
            config['file_type'] = 'excel_multi_tab'
    else:
        # Analyze CSV or single-tab Excel file
        config, debug_entry = analyze_single_file(file_name, file_path, api_key)
        if config:
            # Wrap in standard format for consistency
            config = {
                'file_type': 'csv' if file_type == 'csv' else 'excel_single_tab',
                'config': config
            }

    return config, debug_entry


def analyze_all_files_parallel(uploaded_files, api_key):
    """
    Analyze all uploaded files in parallel using ThreadPoolExecutor.

    V9: Now detects and handles multi-tab Excel files automatically.
    """
    configs = {}
    debug_logs = []

    with ThreadPoolExecutor(max_workers=5) as executor:
        # Submit all tasks (now with file type detection)
        future_to_file = {
            executor.submit(analyze_file_with_detection, file_name, file_path, api_key): file_name
            for file_name, file_path in uploaded_files.items()
        }

        # Collect results as they complete
        for future in as_completed(future_to_file):
            file_name = future_to_file[future]
            try:
                config, debug_entry = future.result()
                if config:
                    configs[file_name] = config
                debug_logs.append(debug_entry)
            except Exception as e:
                debug_logs.append({
                    'file_name': file_name,
                    'timestamp': datetime.now().strftime('%H:%M:%S'),
                    'error': f"Exception: {str(e)}",
                    'success': False
                })

    return configs, debug_logs


def parse_file_with_config(file_path, start_row=0, delimiter=',', num_rows=10):
    """Parse file using the provided configuration."""
    try:
        if str(file_path).lower().endswith(('.xlsx', '.xls')):
            df = pd.read_excel(file_path, header=start_row, nrows=num_rows,
                             dtype=str, keep_default_na=False)
        else:
            df = pd.read_csv(
                file_path,
                sep=delimiter,
                header=start_row,
                nrows=num_rows,
                dtype=str,
                keep_default_na=False,
                encoding='utf-8',
                encoding_errors='ignore',
                on_bad_lines='skip'
            )
        return df
    except Exception as e:
        return None


def calculate_zero_flags(resampled_df, sensor_cols):
    """
    Calculate Zero_Value_Flag for each row based on zero patterns across all sensors.

    Logic:
    - "Clear": No zeros detected in any sensor for this row
    - "Single": At least one sensor has a zero that is NOT consecutive with previous row
    - "Repeated": At least one sensor has 2+ consecutive zeros (current and previous both zero)

    Priority: "Repeated" > "Single" > "Clear"

    V10: Only checks numeric columns (skips text fields like "off"/"on")

    Args:
        resampled_df: The resampled DataFrame
        sensor_cols: List of sensor column names

    Returns:
        List of zero flag values for each row
    """
    zero_flags = []

    for idx in range(len(resampled_df)):
        row_flag = "Clear"

        for sensor in sensor_cols:
            # Skip text columns (only check numeric columns for zeros)
            if not pd.api.types.is_numeric_dtype(resampled_df[sensor]):
                continue

            val_current = resampled_df.loc[idx, sensor]

            # Skip if current value is NaN or not zero
            if pd.isna(val_current) or val_current != 0:
                continue

            # Current value is zero
            if idx > 0:
                val_prev = resampled_df.loc[idx-1, sensor]
                # Check if previous is also zero (repeated)
                if val_prev == 0:
                    row_flag = "Repeated"
                    break  # Highest priority, stop checking
                else:
                    # Single zero (not consecutive)
                    if row_flag == "Clear":
                        row_flag = "Single"
            else:
                # First row, can't be repeated
                if row_flag == "Clear":
                    row_flag = "Single"

        zero_flags.append(row_flag)

    return zero_flags


def resample_to_quarter_hour(combined_df, tolerance_minutes=2):
    """
    Resample combined data to 15-minute intervals with PER-SENSOR nearest-value matching.

    V9 Features:
    - Modified stale flag: 3+ consecutive non-zero identical values (changed from 4+)
    - New Zero_Value_Flag: Clear/Single/Repeated per-sensor tracking
    - Each sensor independently finds its closest value within ±2 minutes
    - Tracks inexact cells for Excel color-coding
    - Consolidated stale flags: single True/False column + comma-separated sensor list

    Args:
        combined_df: The merged dataframe with all sensors
        tolerance_minutes: Window for finding nearest match (±2 minutes default)

    Returns:
        Tuple of (resampled_df, stats_dict, inexact_cells_dict)
    """
    if combined_df is None or combined_df.empty:
        return None, {}, {}

    # Create complete range of 15-minute timestamps
    start_time = combined_df['Date'].min()
    end_time = combined_df['Date'].max()

    # Round start to previous 15-min mark
    start_time = start_time.replace(minute=(start_time.minute // 15) * 15, second=0, microsecond=0)

    # Round end to next 15-min mark
    end_minute = ((end_time.minute // 15) + 1) * 15
    if end_minute >= 60:
        end_time = end_time + timedelta(hours=1)
        end_time = end_time.replace(minute=0, second=0, microsecond=0)
    else:
        end_time = end_time.replace(minute=end_minute, second=0, microsecond=0)

    # Generate 15-minute interval timestamps
    target_timestamps = pd.date_range(start=start_time, end=end_time, freq='15min')

    # Get sensor columns (exclude Date)
    sensor_cols = [col for col in combined_df.columns if col != 'Date']

    # Initialize result lists and inexact cell tracking
    result_data = {'Date': target_timestamps}
    for sensor in sensor_cols:
        result_data[sensor] = []

    inexact_cells = {}  # {row_idx: {sensor: True/False}}
    tolerance = pd.Timedelta(minutes=tolerance_minutes)
    total_inexact = 0

    # For each target timestamp, find nearest value for each sensor independently
    for row_idx, target_time in enumerate(target_timestamps):
        inexact_cells[row_idx] = {}

        # Calculate time differences for all rows
        time_diffs = (combined_df['Date'] - target_time).abs()
        within_window = time_diffs <= tolerance

        # For each sensor, find its nearest value
        for sensor in sensor_cols:
            # Filter to rows within tolerance that have a non-null value for this sensor
            valid_mask = within_window & combined_df[sensor].notna()

            if not valid_mask.any():
                # No value within window - use NULL
                result_data[sensor].append(None)
                inexact_cells[row_idx][sensor] = False
            else:
                # Find the row with the smallest time difference that has this sensor's value
                valid_time_diffs = time_diffs[valid_mask]
                closest_idx = valid_time_diffs.idxmin()

                # Extract the value (preserve 0 as 0, not NULL)
                value = combined_df.loc[closest_idx, sensor]
                result_data[sensor].append(value)

                # Check if the source timestamp is exactly on the quarter-hour mark
                source_time = combined_df.loc[closest_idx, 'Date']
                is_exact = (source_time.minute % 15 == 0) and (source_time.second == 0)
                inexact_cells[row_idx][sensor] = not is_exact

                if not is_exact:
                    total_inexact += 1

    # Create DataFrame
    resampled = pd.DataFrame(result_data)

    # Flag stale data per sensor (temporary for consolidation)
    # V9: Changed to 3+ consecutive non-zero values (was 4+ in V8)
    # V10: Only check numeric columns (skip text fields like "off"/"on")
    stale_per_sensor = {}
    for sensor in sensor_cols:
        # Skip text columns (check if column is numeric)
        if not pd.api.types.is_numeric_dtype(resampled[sensor]):
            # Text column - don't flag for staleness
            stale_per_sensor[sensor] = pd.Series([False] * len(resampled), index=resampled.index)
            continue

        # Skip if value is zero or NaN
        is_non_zero = (resampled[sensor] != 0) & (resampled[sensor].notna())

        # Check if current equals previous 2 values (3 consecutive identical non-zero)
        is_stale = (
            is_non_zero &
            (resampled[sensor] == resampled[sensor].shift(1)) &
            (resampled[sensor] == resampled[sensor].shift(2))
        )
        stale_per_sensor[sensor] = is_stale

    # Consolidate stale flags into two columns
    stale_data_flag = []
    stale_sensors_list = []

    for idx in range(len(resampled)):
        stale_sensors = [sensor for sensor in sensor_cols if stale_per_sensor[sensor].iloc[idx]]

        if stale_sensors:
            stale_data_flag.append(True)
            stale_sensors_list.append(', '.join(stale_sensors))
        else:
            stale_data_flag.append(False)
            stale_sensors_list.append('')

    # Add consolidated stale columns
    resampled['Stale_Data_Flag'] = stale_data_flag
    resampled['Stale_Sensors'] = stale_sensors_list

    # Calculate and add Zero_Value_Flag column (V9 new feature)
    zero_flags = calculate_zero_flags(resampled, sensor_cols)
    resampled['Zero_Value_Flag'] = zero_flags

    # Calculate statistics
    zero_flag_counts = {
        'Clear': zero_flags.count('Clear'),
        'Single': zero_flags.count('Single'),
        'Repeated': zero_flags.count('Repeated')
    }
    stale_counts = {sensor: int(stale_per_sensor[sensor].sum()) for sensor in sensor_cols}
    total_stale_flags = sum(stale_counts.values())

    stats = {
        'total_intervals': len(resampled),
        'total_inexact_cells': int(total_inexact),
        'stale_by_sensor': stale_counts,
        'total_stale_flags': total_stale_flags,
        'rows_with_stale_data': int(sum(stale_data_flag)),
        'zero_flag_counts': zero_flag_counts,
        'date_range': {
            'start': resampled['Date'].min(),
            'end': resampled['Date'].max()
        }
    }

    # Reorder columns: Date, flags, then sensor columns
    flag_cols = ['Stale_Data_Flag', 'Stale_Sensors', 'Zero_Value_Flag']
    sensor_cols_ordered = [col for col in resampled.columns if col not in ['Date'] + flag_cols]
    column_order = ['Date'] + flag_cols + sensor_cols_ordered
    resampled = resampled[column_order]

    return resampled, stats, inexact_cells


def prepare_df_for_display(df):
    """
    Ensure DataFrame is PyArrow-compatible for st.dataframe() display.

    Fixes mixed data types that cause ArrowTypeError during serialization.

    Args:
        df: Pandas DataFrame to prepare

    Returns:
        Copy of DataFrame with corrected types
    """
    if df is None or df.empty:
        return df

    df_copy = df.copy()

    # Convert Date column to datetime if it's object type
    if 'Date' in df_copy.columns:
        if df_copy['Date'].dtype == 'object':
            df_copy['Date'] = pd.to_datetime(df_copy['Date'], errors='coerce')

    # Convert all other object columns to numeric where possible
    for col in df_copy.columns:
        if col not in ['Date', 'Stale_Sensors', 'Zero_Value_Flag', 'Stale_Data_Flag']:
            if df_copy[col].dtype == 'object':
                # Try to convert to numeric
                df_copy[col] = pd.to_numeric(df_copy[col], errors='ignore')

    return df_copy


def build_tab_label(base_name, selected_count, total_count):
    """Build tab label with visual indicators."""
    if selected_count > 0:
        return f"{base_name} ✓ ({selected_count})"
    elif total_count > 0:
        return f"{base_name} ⚠️"
    else:
        return f"{base_name} ❌"


def smart_convert_column(series, threshold=0.8):
    """
    Intelligently convert column to numeric or keep as text.

    Args:
        series: pandas Series to convert
        threshold: Minimum ratio of valid numeric values (default 0.8)

    Returns:
        Converted series (numeric) or original series (text)
    """
    # Try numeric conversion
    numeric_series = pd.to_numeric(series, errors='coerce')

    # Fast path: if ALL are NaN, it's pure text
    if numeric_series.isna().all():
        return series  # Keep original text

    # Check if mostly numeric
    valid_ratio = numeric_series.notna().sum() / len(numeric_series)

    if valid_ratio >= threshold:
        return numeric_series  # SQL-ready numeric
    else:
        return series  # Keep as text


def render_sheet_config_ui(file_name, file_path, sheet_name, config):
    """Render configuration UI for a single Excel sheet."""
    tab_config = config['tabs'][sheet_name]

    # Date column dropdown
    date_col_options = list(range(20))
    date_col_idx = tab_config.get('date_column', 0)

    date_column = st.selectbox(
        "Date Column",
        options=date_col_options,
        index=date_col_idx,
        key=f"date_{file_name}_{sheet_name}"
    )

    # Column selection checkboxes
    st.markdown("**Select Value Columns to Include:**")

    available_cols = tab_config.get('available_columns', [])
    column_names = tab_config.get('column_names', [])
    currently_selected = tab_config.get('selected_columns', available_cols.copy())

    new_selected = []
    for col_idx, col_name in zip(available_cols, column_names):
        is_checked = col_idx in currently_selected
        if st.checkbox(
            f"✓ **{col_name}** (column {col_idx})",
            value=is_checked,
            key=f"col_{file_name}_{sheet_name}_{col_idx}"
        ):
            new_selected.append(col_idx)

    # Update config
    tab_config['date_column'] = date_column
    tab_config['selected_columns'] = new_selected

    # Show preview
    if new_selected:
        st.success(f"✅ {len(new_selected)} column(s) selected from {sheet_name}")

        try:
            df_tab = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                header=tab_config['start_row'],
                nrows=5,
                dtype=str,
                keep_default_na=False
            )
            preview_cols = [df_tab.columns[date_column]] + \
                          [df_tab.columns[i] for i in new_selected if i < len(df_tab.columns)]
            st.dataframe(prepare_df_for_display(df_tab[preview_cols]), height=200)
        except Exception as e:
            st.caption(f"Preview unavailable: {str(e)}")
    else:
        st.warning(f"⚠️ No columns selected from {sheet_name}")


def render_csv_config_ui(file_name, file_path, config):
    """Render configuration UI for CSV or single-tab Excel file."""
    file_type = config.get('file_type', 'csv')
    inner_config = config.get('config', config)  # Handle both formats

    # Configuration inputs
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        start_row = st.number_input(
            "Start Row",
            min_value=0,
            max_value=10,
            value=inner_config.get('start_row', 0),
            key=f"start_{file_name}",
            help="Row where column headers are located"
        )

    with col2:
        delimiter_options = [',', '\t', ';', '|']
        delimiter_labels = ['Comma (,)', 'Tab (\\t)', 'Semicolon (;)', 'Pipe (|)']
        default_delim = inner_config.get('delimiter', ',')
        try:
            delim_idx = delimiter_options.index(default_delim)
        except ValueError:
            delim_idx = 0

        delimiter = st.selectbox(
            "Delimiter",
            options=delimiter_options,
            format_func=lambda x: delimiter_labels[delimiter_options.index(x)],
            index=delim_idx,
            key=f"delim_{file_name}"
        )

    with col3:
        date_column = st.number_input(
            "Date Column",
            min_value=0,
            max_value=20,
            value=inner_config.get('date_column', 0),
            key=f"date_{file_name}",
            help="Column index for date/timestamp (0-based)"
        )

    with col4:
        value_column = st.number_input(
            "Value Column",
            min_value=0,
            max_value=20,
            value=inner_config.get('value_column', 2),
            key=f"value_{file_name}",
            help="Column index for sensor values (0-based)"
        )

    # Sensor name
    sensor_name = st.text_input(
        "Sensor Name",
        value=inner_config.get('sensor_name', Path(file_name).stem),
        key=f"sensor_{file_name}"
    )

    # Update config in session state
    st.session_state.file_configs[file_name] = {
        'file_type': file_type,
        'config': {
            'start_row': start_row,
            'delimiter': delimiter,
            'date_column': date_column,
            'value_column': value_column,
            'sensor_name': sensor_name
        }
    }

    # Show extracted data preview
    st.markdown("**📊 Extracted Data Preview** (what will be used in processing):")
    df_preview = parse_file_with_config(file_path, start_row, delimiter, num_rows=10)
    if df_preview is not None and not df_preview.empty:
        # Create a preview of just the extracted columns
        extracted_preview = pd.DataFrame()

        # Show selected columns info
        col_info1, col_info2 = st.columns(2)
        with col_info1:
            if date_column < len(df_preview.columns):
                date_col_name = df_preview.columns[date_column]
                st.success(f"✅ Date Column [{date_column}]: `{date_col_name}`")
                extracted_preview['Date'] = df_preview.iloc[:, date_column]
            else:
                st.error(f"❌ Date column {date_column} doesn't exist!")

        with col_info2:
            if value_column < len(df_preview.columns):
                value_col_name = df_preview.columns[value_column]
                st.success(f"✅ Value Column [{value_column}]: `{value_col_name}`")
                extracted_preview[sensor_name] = df_preview.iloc[:, value_column]
            else:
                st.error(f"❌ Value column {value_column} doesn't exist!")

        # Show timestamp conversion preview
        if date_column < len(df_preview.columns) and not extracted_preview.empty:
            st.markdown("**🕐 Timestamp Conversion:**")
            # Get first non-null timestamp
            sample_ts = extracted_preview['Date'].dropna().iloc[0] if len(extracted_preview['Date'].dropna()) > 0 else None
            if sample_ts is not None:
                try:
                    original_str = str(sample_ts)
                    normalized = format_timestamp_mdy_hms(original_str)

                    col_ts1, col_ts2, col_ts3 = st.columns([2, 1, 2])
                    with col_ts1:
                        st.text(f"Original: {original_str}")
                    with col_ts2:
                        st.text("→")
                    with col_ts3:
                        st.success(f"Standardized: {normalized}")
                except Exception as e:
                    st.caption(f"Timestamp will be normalized during combine step")

        # Show the extracted data table
        if not extracted_preview.empty:
            st.dataframe(prepare_df_for_display(extracted_preview), height=250)
        else:
            st.warning("No valid columns selected for extraction")
    else:
        st.error("Could not parse file with current settings")

    # Keep raw file preview in collapsible expander
    with st.expander("🔍 View Raw File Data (all columns)", expanded=False):
        if df_preview is not None and not df_preview.empty:
            st.caption("This shows ALL columns from the original file. Only the selected columns above will be used in processing.")
            st.dataframe(prepare_df_for_display(df_preview), height=300)


def extract_multi_tab_data(file_path, file_config):
    """
    Extract data from multi-tab Excel file.

    Args:
        file_path: Path to the Excel file
        file_config: Config dict with structure:
            {
                "file_type": "excel_multi_tab",
                "tabs": {
                    "Tab1": {
                        "start_row": 1,
                        "date_column": 0,
                        "available_columns": [2, 3, 4],
                        "column_names": ["Col1", "Col2", "Col3"],
                        "selected_columns": [2, 3]  # User-selected subset
                    }
                }
            }

    Returns:
        List of DataFrames, each with [Date, SensorName] columns
    """
    all_dataframes = []

    for tab_name, tab_config in file_config['tabs'].items():
        try:
            # Read the tab
            df = pd.read_excel(
                file_path,
                sheet_name=tab_name,
                header=tab_config['start_row'],
                dtype=str,
                keep_default_na=False
            )

            # Extract Date column
            date_col_idx = tab_config['date_column']
            date_series = df.iloc[:, date_col_idx]

            # Build dictionary with all selected columns for this tab
            selected_data = {}
            selected_data['Date'] = date_series

            # Extract selected value columns
            selected_indices = tab_config['selected_columns']
            column_names = tab_config['column_names']

            for col_idx in selected_indices:
                # Find position in available_columns to get correct name
                try:
                    name_idx = tab_config['available_columns'].index(col_idx)
                    col_name = column_names[name_idx]
                except (ValueError, IndexError):
                    col_name = f"Column_{col_idx}"

                # Create final column name: TabName ColumnName
                final_name = f"{tab_name} {col_name}"

                # Extract and intelligently convert value column (preserves text, converts numeric)
                value_series = smart_convert_column(df.iloc[:, col_idx], threshold=0.8)
                selected_data[final_name] = value_series

            # Create single DataFrame per tab with all selected columns
            tab_df = pd.DataFrame(selected_data)

            # Normalize timestamps ONCE per tab (not per column)
            tab_df['Date'] = tab_df['Date'].apply(
                lambda x: format_timestamp_mdy_hms(str(x)) if pd.notna(x) else None
            )

            # Convert Date to datetime
            tab_df['Date'] = pd.to_datetime(tab_df['Date'], format='%m/%d/%Y %H:%M:%S', errors='coerce')

            # Drop rows with invalid dates
            tab_df = tab_df.dropna(subset=['Date'])

            if not tab_df.empty:
                all_dataframes.append(tab_df)

        except Exception as e:
            print(f"Error extracting data from tab {tab_name}: {str(e)}")
            continue

    return all_dataframes


def detect_percentage_columns(file_path, sheet_name, col_indices):
    """
    Check if columns have percentage formatting in source Excel file.

    Args:
        file_path: Path to Excel file
        sheet_name: Name of the sheet/tab
        col_indices: List of column indices to check

    Returns:
        List of column indices that have percentage formatting
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=False)

        if sheet_name not in wb.sheetnames:
            return []

        ws = wb[sheet_name]
        percentage_cols = []

        for col_idx in col_indices:
            # Check first data row (row 2 assuming row 1 is header)
            cell = ws.cell(row=2, column=col_idx + 1)  # Excel is 1-indexed
            if cell.number_format and '%' in cell.number_format:
                percentage_cols.append(col_idx)

        wb.close()
        return percentage_cols

    except Exception as e:
        print(f"Error detecting percentage columns: {str(e)}")
        return []


def export_to_excel(resampled_df, inexact_cells, output_path):
    """
    Export resampled data to Excel with color-coded quality indicators.

    V9 Features:
    - Inexact cells are highlighted in yellow
    - Stale_Data_Flag column highlights True values in light red
    - Zero_Value_Flag column included (Clear/Single/Repeated)
    - Date column formatted as text to preserve formatting

    Args:
        resampled_df: The resampled DataFrame
        inexact_cells: Dictionary mapping {row_idx: {sensor: True/False}}
        output_path: Path to save the Excel file

    Returns:
        True if successful, False otherwise
    """
    try:
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Resampled Data"

        # Define color fills
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow for inexact
        red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light red for stale

        # Get sensor columns (exclude Date and flag columns)
        sensor_cols = [col for col in resampled_df.columns
                      if col not in ['Date', 'Stale_Data_Flag', 'Stale_Sensors', 'Zero_Value_Flag']]

        # Create column index mapping
        col_indices = {col: idx + 1 for idx, col in enumerate(resampled_df.columns)}

        # Write header row
        for col_idx, col_name in enumerate(resampled_df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True)

        # Write data rows and apply coloring
        for row_idx, row_data in enumerate(resampled_df.itertuples(index=False), start=2):
            data_row_idx = row_idx - 2  # Adjust for 0-based indexing in inexact_cells

            for col_idx, (col_name, value) in enumerate(zip(resampled_df.columns, row_data), start=1):
                cell = ws.cell(row=row_idx, column=col_idx)

                # Handle date formatting
                if col_name == 'Date':
                    if pd.notna(value):
                        cell.value = value.strftime('%m/%d/%Y %H:%M:%S')
                    else:
                        cell.value = None
                else:
                    cell.value = value

                # Color sensor cells if inexact
                if col_name in sensor_cols and data_row_idx in inexact_cells:
                    if inexact_cells[data_row_idx].get(col_name, False):
                        cell.fill = yellow_fill

                # Color Stale_Data_Flag if True
                if col_name == 'Stale_Data_Flag' and value == True:
                    cell.fill = red_fill

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save workbook
        wb.save(output_path)
        return True

    except Exception as e:
        print(f"Error exporting to Excel: {str(e)}")
        return False


def archive_uploaded_files(uploaded_files, archive_path):
    """
    Copy original files to archive directory for safekeeping.

    Args:
        uploaded_files: Dictionary of {filename: filepath}
        archive_path: Destination directory path

    Returns:
        List of archived file paths
    """
    try:
        # Create archive directory
        archive_dir = Path(archive_path)
        archive_dir.mkdir(parents=True, exist_ok=True)

        # Copy each file
        archived_files = []
        for file_name, file_path in uploaded_files.items():
            dest_path = archive_dir / file_name
            shutil.copy2(file_path, dest_path)
            archived_files.append(str(dest_path))

        return archived_files

    except Exception as e:
        raise Exception(f"Archiving failed: {str(e)}")


def main():
    """Main application interface."""

    # Subheader below logo
    st.markdown("**Tab-Based UI** • **Smart Data Types** • **Enhanced Flagging** • **Multi-Tab Excel** • **AI Analysis**")
    st.markdown("---")

    # ========== STEP 1: UPLOAD ==========
    st.header("Step 1: Upload Files & Archive Settings")

    # Building Name Input (Required for archiving)
    building_name = st.text_input(
        "Building Name",
        value=st.session_state.get('building_name', ''),
        key="building_name",
        placeholder="e.g., Gotham Tower",
        help="Used for organizing archived files"
    )

    # Archive Location Selection
    st.markdown("### 📦 Archive Settings")

    use_custom_path = st.checkbox(
        "Use custom archive location",
        value=st.session_state.get('use_custom_archive', False),
        key="use_custom_archive",
        help="Check this to manually specify where to save archived files"
    )

    if use_custom_path:
        # Manual filepath entry
        archive_path = st.text_input(
            "Custom Archive Folder Path",
            value=st.session_state.get('archive_path', ''),
            key="archive_path",
            placeholder="e.g., C:/My Files/Archive or D:/Backups",
            help="Enter the full path where you want to save the original files"
        )
        st.caption("💡 Enter a custom path above (e.g., `C:/MyFolder/Archive` or `D:/Backups/BuildingData`)")
    else:
        # Default archive structure: archive/[Building Name]/
        if building_name and building_name.strip():
            default_archive = f"archive/{building_name}"
        else:
            default_archive = "archive/Unnamed"

        archive_path = default_archive
        st.session_state.archive_path = default_archive

        st.info(f"📁 **Default Archive Location:** `{default_archive}/`")
        st.caption("Files will be saved to the archive folder in your project directory")

    st.caption("📦 **All uploaded files will be archived** before processing")
    st.markdown("")  # Spacing

    # File Uploader
    uploaded_files = st.file_uploader(
        "📁 Upload your sensor files",
        type=['xlsx', 'csv', 'xls'],
        accept_multiple_files=True,
        help="Select all files (10-90 files)"
    )

    if uploaded_files:
        # Save files to temp directory
        temp_dir = Path("temp")
        temp_dir.mkdir(exist_ok=True)

        for uploaded_file in uploaded_files:
            file_path = temp_dir / uploaded_file.name
            if uploaded_file.name not in st.session_state.uploaded_files:
                with open(file_path, "wb") as f:
                    f.write(uploaded_file.getbuffer())
                st.session_state.uploaded_files[uploaded_file.name] = str(file_path)

        # Always archive files to the specified path
        archive_path = st.session_state.get('archive_path', '')

        if archive_path and archive_path.strip():
            try:
                archived_files = archive_uploaded_files(
                    st.session_state.uploaded_files,
                    archive_path
                )
                st.success(f"✅ {len(uploaded_files)} files uploaded and archived to:")
                st.code(archive_path, language="")
            except Exception as e:
                st.error(f"❌ Archiving failed: {e}")
                st.warning("Files are saved in temp storage but not permanently archived.")
                st.info("Please check the archive path and try re-uploading, or proceed with processing anyway.")
        else:
            st.warning("⚠️ Archive path is empty - files uploaded to temp storage only")
            st.info("Files will be processed but not permanently archived")

        st.markdown("---")

        # ========== STEP 2: AI ANALYSIS ==========
        st.header("Step 2: AI Analysis")

        # Check for API key
        api_key = os.getenv('CLAUDE_API_KEY')
        if not api_key:
            st.error("⚠️ CLAUDE_API_KEY not found in .env file. Please add your API key.")
            st.code("CLAUDE_API_KEY=sk-ant-...", language="bash")
        else:
            col1, col2 = st.columns([2, 1])

            with col1:
                st.info("💡 Click below to analyze ALL files in parallel with Claude AI")

            with col2:
                if st.button("🤖 Analyze All Files", type="primary"):
                    with st.spinner(f"Analyzing {len(st.session_state.uploaded_files)} files in parallel..."):
                        progress_bar = st.progress(0)

                        # Run parallel analysis
                        configs, debug_logs = analyze_all_files_parallel(
                            st.session_state.uploaded_files,
                            api_key
                        )

                        progress_bar.progress(100)

                        # Store results
                        st.session_state.file_configs = configs
                        st.session_state.ai_debug_log = debug_logs
                        st.session_state.ai_analysis_complete = True

                        # Show summary
                        success_count = sum(1 for log in debug_logs if log.get('success', False))
                        st.success(f"✅ Analysis complete! {success_count}/{len(debug_logs)} files configured successfully")
                        st.rerun()

            # Show analysis results if complete
            if st.session_state.ai_analysis_complete and st.session_state.file_configs:
                st.markdown("---")
                st.subheader("Analysis Results")

                # Summary metrics
                total_files = len(st.session_state.uploaded_files)
                configured_files = len(st.session_state.file_configs)
                success_rate = (configured_files / total_files * 100) if total_files > 0 else 0

                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Total Files", total_files)
                with col2:
                    st.metric("Configured", configured_files)
                with col3:
                    st.metric("Success Rate", f"{success_rate:.0f}%")

        # ========== STEP 3: REVIEW & EDIT CONFIGS ==========
        if st.session_state.file_configs:
            st.markdown("---")
            st.header("Step 3: Review & Edit Configurations")

            st.info("💡 Review AI-detected settings. Click tabs to configure each file.")

            # Build file-level tabs
            file_tab_labels = []
            file_names_ordered = list(st.session_state.uploaded_files.keys())

            for file_name in file_names_ordered:
                config = st.session_state.file_configs.get(file_name, {})
                file_type = config.get('file_type', 'csv')

                # Calculate total selected columns
                if file_type == 'excel_multi_tab':
                    total_selected = sum(
                        len(tab_cfg.get('selected_columns', []))
                        for tab_cfg in config['tabs'].values()
                    )
                    total_available = sum(
                        len(tab_cfg.get('available_columns', []))
                        for tab_cfg in config['tabs'].values()
                    )
                else:
                    total_selected = 1
                    total_available = 1

                label = build_tab_label(file_name, total_selected, total_available)
                file_tab_labels.append(label)

            # Render file-level tabs
            file_tabs = st.tabs(file_tab_labels)

            for file_tab, file_name in zip(file_tabs, file_names_ordered):
                with file_tab:
                    config = st.session_state.file_configs.get(file_name, {})
                    file_type = config.get('file_type', 'csv')
                    file_path = st.session_state.uploaded_files[file_name]

                    if file_type == 'excel_multi_tab':
                        st.info(f"📑 **Multi-Tab Excel** - {len(config['tabs'])} sheets")

                        # Build sheet-level tabs
                        sheet_tab_labels = []
                        sheet_names_ordered = list(config['tabs'].keys())

                        for sheet_name in sheet_names_ordered:
                            tab_cfg = config['tabs'][sheet_name]
                            selected_count = len(tab_cfg.get('selected_columns', []))
                            available_count = len(tab_cfg.get('available_columns', []))
                            label = build_tab_label(sheet_name, selected_count, available_count)
                            sheet_tab_labels.append(label)

                        # Render sheet-level tabs
                        sheet_tabs = st.tabs(sheet_tab_labels)

                        for sheet_tab, sheet_name in zip(sheet_tabs, sheet_names_ordered):
                            with sheet_tab:
                                render_sheet_config_ui(file_name, file_path, sheet_name, config)

                        # Update config
                        st.session_state.file_configs[file_name] = config

                    else:
                        # CSV/Single-tab: Render config directly
                        render_csv_config_ui(file_name, file_path, config)

        # ========== STEP 4: COMBINE DATA (RAW MERGE) ==========
        if st.session_state.file_configs and len(st.session_state.file_configs) == len(st.session_state.uploaded_files):
            st.markdown("---")
            st.header("Step 4: Combine All Data (Raw Merge)")

            # Show timestamp normalization preview
            st.info("🕐 **Timestamp Standardization**: All timestamps will be normalized to **MM/DD/YYYY HH:MM:SS** format")

            with st.expander("📅 Preview Timestamp Conversion", expanded=False):
                st.markdown("**Sample conversions from your uploaded files:**")

                # Show a few examples from each file (skip multi-tab Excel for simplicity)
                for file_name, config in list(st.session_state.file_configs.items())[:3]:  # Show first 3 files
                    file_type = config.get('file_type', 'csv')

                    # Skip multi-tab Excel files in preview (too complex)
                    if file_type == 'excel_multi_tab':
                        st.info(f"**{file_name}**: Multi-tab Excel - timestamps will be normalized during combine")
                        continue

                    # Get inner config for CSV/single-tab files
                    inner_config = config.get('config', config)

                    file_path = st.session_state.uploaded_files[file_name]
                    df_sample = parse_file_with_config(
                        file_path,
                        inner_config.get('start_row', 0),
                        inner_config.get('delimiter', ','),
                        num_rows=3
                    )

                    if df_sample is not None and inner_config.get('date_column', 0) < len(df_sample.columns):
                        st.markdown(f"**{file_name}:**")

                        # Get first non-null timestamp
                        date_col_name = df_sample.columns[inner_config['date_column']]
                        sample_timestamps = df_sample[date_col_name].dropna().head(2)

                        for original_ts in sample_timestamps:
                            try:
                                original_str = str(original_ts)
                                detected_format = detect_timestamp_format(original_str)
                                normalized = format_timestamp_mdy_hms(original_str)

                                col1, col2, col3 = st.columns([2, 1, 2])
                                with col1:
                                    st.text(f"Original: {original_str}")
                                with col2:
                                    st.text("→")
                                with col3:
                                    st.success(f"Standardized: {normalized}")

                                st.caption(f"Detected: {detected_format}")
                                st.markdown("---")
                            except Exception as e:
                                st.warning(f"Could not parse: {original_ts} ({str(e)})")

            if st.button("🔗 Combine All Files", type="primary"):
                with st.spinner("Combining all files..."):
                    progress_bar = st.progress(0)
                    status_text = st.empty()

                    try:
                        loaded_dfs = []
                        total_files = len(st.session_state.file_configs)

                        for idx, (file_name, config) in enumerate(st.session_state.file_configs.items()):
                            status_text.text(f"Processing {file_name}... ({idx + 1}/{total_files})")
                            progress_bar.progress((idx + 1) / total_files)

                            file_path = st.session_state.uploaded_files[file_name]
                            file_type = config.get('file_type', 'csv')

                            # ===== MULTI-TAB EXCEL FILE =====
                            if file_type == 'excel_multi_tab':
                                # Extract data from all selected columns in all tabs
                                multi_tab_dfs = extract_multi_tab_data(file_path, config)
                                loaded_dfs.extend(multi_tab_dfs)

                            # ===== CSV OR SINGLE-TAB EXCEL FILE =====
                            else:
                                inner_config = config.get('config', config)

                                # Read full file
                                if file_path.lower().endswith(('.xlsx', '.xls')):
                                    df_full = pd.read_excel(
                                        file_path,
                                        header=inner_config['start_row'],
                                        dtype=str,
                                        keep_default_na=False
                                    )
                                else:
                                    df_full = pd.read_csv(
                                        file_path,
                                        sep=inner_config['delimiter'],
                                        header=inner_config['start_row'],
                                        dtype=str,
                                        keep_default_na=False,
                                        encoding='utf-8',
                                        encoding_errors='ignore',
                                        on_bad_lines='skip'
                                    )

                                # Extract required columns
                                df_clean = pd.DataFrame()

                                # Get date column
                                if inner_config['date_column'] < len(df_full.columns):
                                    date_col_name = df_full.columns[inner_config['date_column']]
                                    df_clean['Date'] = df_full[date_col_name]

                                # Get value column
                                if inner_config['value_column'] < len(df_full.columns):
                                    value_col_name = df_full.columns[inner_config['value_column']]
                                    df_clean[inner_config['sensor_name']] = df_full[value_col_name]

                                # Normalize and standardize timestamps
                                if 'Date' in df_clean.columns:
                                    # Apply timestamp normalization
                                    normalized_dates = []
                                    for ts_val in df_clean['Date']:
                                        try:
                                            # Normalize to MM/DD/YYYY HH:MM:SS format
                                            normalized = format_timestamp_mdy_hms(str(ts_val))
                                            # Parse back to datetime for consistency
                                            normalized_dates.append(pd.to_datetime(normalized, format='%m/%d/%Y %H:%M:%S'))
                                        except Exception:
                                            # If normalization fails, try pandas default parsing
                                            try:
                                                normalized_dates.append(pd.to_datetime(ts_val, format='mixed', errors='coerce'))
                                            except:
                                                normalized_dates.append(pd.NaT)

                                    df_clean['Date'] = normalized_dates
                                    df_clean = df_clean.dropna(subset=['Date'])

                                    # Remove timezone if present (we standardized to America/New_York)
                                    if pd.api.types.is_datetime64tz_dtype(df_clean['Date']):
                                        df_clean['Date'] = df_clean['Date'].dt.tz_localize(None)

                                    loaded_dfs.append(df_clean)

                        # Combine all dataframes using reduce() for cleaner merging
                        if loaded_dfs:
                            from functools import reduce

                            combined = reduce(
                                lambda left, right: pd.merge(left, right, on='Date', how='outer'),
                                loaded_dfs
                            )

                            # Sort by date and remove exact duplicates
                            combined = combined.sort_values('Date').reset_index(drop=True)
                            combined = combined.drop_duplicates()

                            st.session_state.combined_df = combined

                            # Clear progress indicators
                            progress_bar.empty()
                            status_text.empty()

                            st.success(f"✅ Successfully combined {total_files} files!")
                            st.balloons()

                            # Show summary metrics
                            col1, col2, col3 = st.columns(3)
                            with col1:
                                st.metric("Total Rows", f"{len(combined):,}")
                            with col2:
                                num_sensors = len([c for c in combined.columns if c != 'Date'])
                                st.metric("Sensors", num_sensors)
                            with col3:
                                date_range = f"{combined['Date'].min().strftime('%m/%d/%Y')} - {combined['Date'].max().strftime('%m/%d/%Y')}"
                                st.metric("Date Range", date_range)

                            # Show preview
                            st.markdown("### Raw Merged Data Preview")
                            st.dataframe(prepare_df_for_display(combined.head(20)), height=400)

                            st.info("💡 **This is raw merged data.** Proceed to Step 5 to resample to 15-minute intervals with quality flags.")

                    except Exception as e:
                        # Clear progress indicators
                        progress_bar.empty()
                        status_text.empty()

                        st.error(f"❌ Error combining files: {str(e)}")
                        st.exception(e)

        # ========== DOWNLOAD RAW MERGED CSV ==========
        if st.session_state.combined_df is not None and st.session_state.resampled_df is None:
            st.markdown("---")
            st.subheader("📥 Download Raw Merged Data (Optional)")

            st.info("💾 You can download the raw merged data before resampling. This preserves all original timestamps.")

            col1, col2 = st.columns([2, 1])
            with col1:
                raw_filename = st.text_input(
                    "Raw merged filename",
                    value=f"raw_merged_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    key="raw_filename"
                )

            with col2:
                st.markdown("###")
                # Generate download button
                combined = st.session_state.combined_df
                combined_export = combined.copy()
                if 'Date' in combined_export.columns:
                    combined_export['Date'] = combined_export['Date'].dt.strftime('%m/%d/%Y %H:%M:%S')

                csv_data = combined_export.to_csv(index=False)
                st.download_button(
                    label="⬇️ Download Raw Merged CSV",
                    data=csv_data,
                    file_name=raw_filename,
                    mime='text/csv'
                )

        # ========== STEP 5: RESAMPLE TO 15-MINUTE INTERVALS ==========
        if st.session_state.combined_df is not None:
            st.markdown("---")
            st.header("Step 5: Resample to Quarter-Hour Intervals")

            st.info("""
            🕐 **Quarter-Hour Resampling Process (Per-Sensor Matching) - V8:**
            - Creates complete 15-minute interval timestamps (:00, :15, :30, :45)
            - **Each sensor independently** finds its closest value within ±2 minute tolerance
            - Different sensors can pull from different source timestamps
            - **Inexact cells will be color-coded yellow in Excel output**
            - Consolidated stale flags: one column + comma-separated sensor list
            - Uses NULL when no value exists within ±2 minute window
            """)

            if st.button("⏱️ Resample to 15-Minute Intervals", type="primary"):
                with st.spinner("Resampling to quarter-hour intervals..."):
                    try:
                        resampled, stats, inexact_cells = resample_to_quarter_hour(st.session_state.combined_df)

                        if resampled is not None:
                            st.session_state.resampled_df = resampled
                            st.session_state.resampling_stats = stats
                            st.session_state.inexact_cells = inexact_cells

                            st.success("✅ Successfully resampled to 15-minute intervals!")
                            st.balloons()

                            # Show statistics
                            st.markdown("### Resampling Statistics")

                            col1, col2, col3, col4 = st.columns(4)
                            with col1:
                                st.metric("Total Intervals", f"{stats['total_intervals']:,}")
                            with col2:
                                st.metric("Inexact Cells", f"{stats['total_inexact_cells']:,}")
                            with col3:
                                st.metric("Rows with Stale Data", stats['rows_with_stale_data'])
                            with col4:
                                st.metric("Total Stale Flags", stats['total_stale_flags'])

                            # Stale data by sensor
                            if stats['stale_by_sensor']:
                                st.markdown("#### Stale Data by Sensor")
                                stale_df = pd.DataFrame([
                                    {'Sensor': sensor, 'Stale Count': count}
                                    for sensor, count in stats['stale_by_sensor'].items()
                                ]).sort_values('Stale Count', ascending=False)
                                st.dataframe(stale_df)

                            # Show preview
                            st.markdown("### Resampled Data Preview")
                            st.dataframe(prepare_df_for_display(resampled.head(20)), height=400)

                        else:
                            st.error("❌ Failed to resample data")

                    except Exception as e:
                        st.error(f"❌ Error during resampling: {str(e)}")
                        st.exception(e)

        # ========== STEP 6: EXPORT RESAMPLED DATA ==========
        if st.session_state.resampled_df is not None:
            st.markdown("---")
            st.header("Step 6: Export Resampled Data")

            resampled = st.session_state.resampled_df

            st.info("""
            📊 **Excel Output Features:**
            - **Yellow cells**: Inexact matches (not exactly on quarter-hour mark)
            - **Light red cells**: Rows with stale data
            - **Stale_Sensors column**: Comma-separated list of sensors with stale data
            """)

            col1, col2 = st.columns([2, 1])
            with col1:
                output_filename = st.text_input(
                    "Output filename",
                    value=f"resampled_15min_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                )

            with col2:
                st.markdown("###")
                if st.button("📥 Generate Excel File", type="primary"):
                    try:
                        output_dir = Path("output")
                        output_dir.mkdir(exist_ok=True)
                        output_path = output_dir / output_filename

                        # Export to Excel with color coding
                        success = export_to_excel(
                            resampled,
                            st.session_state.inexact_cells,
                            output_path
                        )

                        if success:
                            st.success(f"✅ Excel file saved to: {output_path}")

                            # Download button
                            with open(output_path, 'rb') as f:
                                st.download_button(
                                    label="⬇️ Download Excel File",
                                    data=f,
                                    file_name=output_filename,
                                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                                )
                        else:
                            st.error("❌ Failed to create Excel file")

                    except Exception as e:
                        st.error(f"❌ Error creating file: {str(e)}")

            # Export statistics
            with st.expander("📊 Export Statistics"):
                st.markdown("### Data Summary")

                # Get sensor columns (exclude Date and flag columns)
                sensor_cols = [c for c in resampled.columns
                              if c not in ['Date', 'Stale_Data_Flag', 'Stale_Sensors']]

                if sensor_cols:
                    st.write(resampled[sensor_cols].describe())

                st.markdown("### Quality Flags Summary")

                # Inexact cells (calculated from metadata)
                if st.session_state.inexact_cells:
                    total_inexact = sum(
                        sum(1 for is_inexact in row_data.values() if is_inexact)
                        for row_data in st.session_state.inexact_cells.values()
                    )
                    st.write(f"**Total Inexact Cells**: {total_inexact:,}")
                    st.caption("These cells are highlighted in yellow in the Excel file")

                # Stale data flags (consolidated)
                if 'Stale_Data_Flag' in resampled.columns:
                    stale_rows = resampled['Stale_Data_Flag'].sum()
                    stale_pct = (stale_rows / len(resampled)) * 100
                    st.write(f"**Rows with Stale Data**: {stale_rows:,} ({stale_pct:.1f}%)")
                    st.caption("These rows are highlighted in light red in the Excel file")

                    # Show detailed breakdown from stats
                    if st.session_state.resampling_stats and 'stale_by_sensor' in st.session_state.resampling_stats:
                        st.write("**Stale Data by Sensor:**")
                        stale_by_sensor = st.session_state.resampling_stats['stale_by_sensor']
                        for sensor, count in sorted(stale_by_sensor.items(), key=lambda x: x[1], reverse=True):
                            if count > 0:
                                st.write(f"- {sensor}: {count:,}")

                st.markdown("### Missing Data Report")
                missing_data = resampled[sensor_cols].isnull().sum()
                missing_data = missing_data[missing_data > 0]
                if not missing_data.empty:
                    st.write(missing_data)
                else:
                    st.success("No missing data in sensor columns!")

        # ========== AI DEBUG PANEL (ALWAYS VISIBLE) ==========
        st.markdown("---")
        st.header("🔍 AI Debug Panel")

        if not st.session_state.ai_debug_log:
            st.info("No AI analysis yet. Upload files and click 'Analyze All Files' to see debug information.")
        else:
            # Summary
            total_calls = len(st.session_state.ai_debug_log)
            successful_calls = sum(1 for log in st.session_state.ai_debug_log if log.get('success', False))

            col1, col2, col3 = st.columns([2, 2, 1])
            with col1:
                st.metric("Total API Calls", total_calls)
            with col2:
                st.metric("Successful", f"{successful_calls}/{total_calls}")
            with col3:
                if st.button("🗑️ Clear Log"):
                    st.session_state.ai_debug_log = []
                    st.rerun()

            st.markdown("---")

            # Display each debug entry
            for idx, entry in enumerate(st.session_state.ai_debug_log):
                with st.expander(f"📄 {entry['file_name']} - {'✅ Success' if entry.get('success') else '❌ Failed'}", expanded=False):
                    col1, col2 = st.columns([1, 1])

                    with col1:
                        st.markdown("**⏰ Timestamp**")
                        st.text(entry['timestamp'])

                        if entry.get('request'):
                            st.markdown("**📤 Request Details**")
                            st.json({
                                'model': entry['request']['model'],
                                'max_tokens': entry['request']['max_tokens'],
                                'temperature': entry['request']['temperature'],
                                'prompt_length': entry['request']['prompt_length']
                            })

                    with col2:
                        if entry.get('response'):
                            st.markdown("**📥 Response**")
                            if 'parsed_json' in entry['response']:
                                st.json(entry['response']['parsed_json'])
                            else:
                                st.code(entry['response'].get('raw_text', 'No response')[:500])

                        if entry.get('error'):
                            st.markdown("**❌ Error**")
                            st.error(entry['error'])

                    # Full prompt and response
                    if entry.get('request'):
                        with st.expander("View Full Prompt"):
                            st.code(entry['request'].get('prompt_preview', ''), language='text')

                            st.markdown("**First 5 Lines Sent:**")
                            for line_idx, line in enumerate(entry['request'].get('raw_text_lines', [])):
                                st.text(f"{line_idx}: {line[:100]}")

                    if entry.get('response'):
                        with st.expander("View Full Response"):
                            st.code(entry['response'].get('raw_text', ''))


if __name__ == "__main__":
    main()
